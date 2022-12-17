import requests
import json
from openpyxl import Workbook
from requests.structures import CaseInsensitiveDict
from random import randint


product_temp = []
products_id = int()
products_category1 = str()
products_category2 = str()
products_name = str()
products_url = str()
products_seller = str()
products_seller_id = int()
products_price = int()
products_pick = bool()
products_delivery_day = str()
products_delivery_price = int()
products_position = int()
products_comparison = str()
products_categories = list()
category = str()
temp2 = list()
links_list = list()


def get_a_link():
    global category
    text = url.get().split("/")
    category = text[-2]
    merchantid = merchant.get()
    # temp = requests.get(f"https://www.toqaev.tk/getCategory.php?cat={category1}").json()
    # global products_category1
    # products_category1 = category1
    # return temp
    choose_category2(text[-2], merchantid)


def choose_category2(category, merchantid, products=list()):
    global product_temp
    # root.destroy()
    page = 0
    req = f"https://kaspi.kz/yml/product-view/pl/filters?q=:category:{category}&text&all&sort=relevance&ui=d&i=-1&page={page}"
    headers = CaseInsensitiveDict()
    headers["Referer"] = f"https://kaspi.kz/shop/c/{category}/"
    headers["Accept"] = "application/json, text/*"
    headers["User-Agent"] = agent
    temp2 = requests.get(req, headers=headers).json()
    some = int(temp2["data"]["total"])
    links_list = [[*range(0, some, 8)], [*range(1, some, 8)], [*range(2, some, 8)], [*range(3, some, 8)],
                  [*range(4, some, 8)], [*range(5, some, 8)], [*range(6, some, 8)], [*range(7, some, 8)]]

    print(some)
    try:
        # p = Pool(processes=8)
        print(links_list)
        # for i in range(len(links_list[0])):
        #   links_list2 = [[links_list[0][i]], [links_list[1][i]], [links_list[2][i]], [links_list[3][i]]
        #                     , [links_list[4][i]], [links_list[5][i]], [links_list[6][i]], [links_list[7][i]]]
        # print(links_list2)

        threads = list()
        for index in range(8):
            logging.info("Main    : create and start thread %d.", index)
            x = threading.Thread(target=parse, args=(links_list[index], product_temp,) )
            threads.append(x)
            x.start()

        for index, thread in enumerate(threads):
            logging.info("Main    : before joining thread %d.", index)
            # product_temp.append(thread.get())
            thread.join()
            logging.info("Main    : thread %d done", index)
        # product_temp.append(p.map(parse, links_list))
        # with Pool(5) as p:
        #    product_temp.append(p.map(parse, links_list))
        # print(products)


    # except Exception as ex:
    #   print(ex)
    # #  pass
    finally:
        print("FROM NOW ON")
        print(len(product_temp))
        product_temp = sorted(product_temp, key=lambda d: d['id'])
        print(product_temp[0:10])
        # product_temp = [result[0] for result in results]
        # for i in product_temp:
        #     for g in i:
        #         for t in g:
        #             products.append(t)
        products = list(set(product_temp))
        # for d in product_temp:
        #     count = 0
        #     for y in product_temp:
        #         # if d["seller_id"] == "Mebelskz":
        #         # print("________________")
        #         # print(product_temp[d])
        #         # print(product_temp[y])
        #         # print("________________")
        #         if d["id"] == y["id"] and d["seller_id"] == y["seller_id"]:
        #             count= count + 1
        #             print(f'{d["seller"]} {d["name"]}')
        #             print(count)
        #             if count > 1:
        #                 break
        #     if count < 2:
        #         products.append(d)
        #     # seen.append(product_temp[d])
        #     else:
        #         pass
        print(len(products))

        new_list = sorted(products, key=lambda d: d['id'])
        print(len(new_list))
        # print(new_list)
        products = []
        for i in range(len(new_list)):
            if i == 0:
                products.append([new_list[i]])
            else:
                for g in range(len(products)):
                    # print(products)
                    if products[g][0]["id"] == new_list[i]["id"]:
                        products[g].append(new_list[i])
                        break
                products.append([new_list[i]])

        save = 0
        for i in range(len(products)):
            for j in products[i]:
                if save == j["id"]:
                    products[i].remove(j)
                    print("yep")
                if j["seller_id"] == merchantid:
                    save = j["id"]
                    print("uhh")
        print(products[2])
        # print("HERE")
        # print("HERE")
        # print(products)
        # print("HERE")
        compare(merchantid, products)
        # print(products_final[41])


gandom = True


def parse(pages, product_temp):
    print(category)
    global gandom
    for page in pages:
        if gandom:
            pass
        else:
            return
        req = f"https://kaspi.kz/yml/product-view/pl/filters?q=:category:{category}&text&all&sort=relevance&ui=d&i=-1&page={page}"

        headers = CaseInsensitiveDict()
        headers["Referer"] = f"https://kaspi.kz/shop/c/{category}/"
        headers["Accept"] = "application/json, text/*"
        headers["User-Agent"] = agent
        temp2 = requests.get(req, headers=headers).json()
        if len(temp2["data"]["cards"]) < 12:
            print(f"The end is near {len(temp2['data']['cards'])}")
            gandom = False
            for product in temp2["data"]["cards"]:  # по товарам прохоl
                products_id = product['id']
                products_name = product['title']
                products_url = product['shopLink']
                products_categories = product['category']
                req2 = f"https://kaspi.kz/yml/offer-view/offers/{product['id']}"
                offer_headers = CaseInsensitiveDict()
                offer_headers["Accept"] = "application/json, text/*"
                offer_headers["Referer"] = product["shopLink"]
                offer_headers["Origin"] = "https://kaspi.kz"
                offer_headers["Content-Type"] = "application/json"
                offer_headers["User-Agent"] = agent
                offer_count = 5
                offer_data = json.dumps(
                    {"cityId": "351010000", "id": str(products_id), "merchantUID": "", "limit": offer_count,
                     "page": 0, "sort": True, "baseProductCodes": [], "groups": [], "installationId": "-1"})
                temp3 = requests.post(req2, headers=offer_headers, data=offer_data).json()
                offer_count = temp3['total']
                offer_data = json.dumps(
                    {"cityId": "351010000", "id": str(products_id), "merchantUID": "", "limit": offer_count,
                     "page": 0, "sort": True, "baseProductCodes": [], "groups": [], "installationId": "-1"})
                temp3 = requests.post(req2, headers=offer_headers, data=offer_data).json()
                count = 1
                for offer in temp3["offers"]:
                    products_seller = offer["merchantName"]
                    products_seller_id = offer["merchantId"]
                    products_price = offer["price"]
                    months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь",
                              "Октябрь", "Ноябрь", "Декабрь"]
                    try:
                        products_pick = f"{offer['pickup'][8:10]} {months[int(offer['delivery'][5:7])]}"
                        products_pick = "Есть"
                    except:
                        products_pick = "Нет"
                    try:
                        products_delivery_day = f"{offer['delivery'][8:10]} {months[int(offer['delivery'][5:7])]}"
                        products_delivery_price = "Бесплатно"
                    except:
                        products_delivery_day = "Нет"
                        products_delivery_price = "Бесплатно"
                    products_position = count
                    count += 1
                    product_temp.append({
                        "id": int(products_id),
                        "category1": str(products_categories[0]),
                        "category3": str(products_categories[2]),
                        "category2": str(products_categories[1]),
                        "name": str(products_name),
                        "url": str(products_url),
                        "seller": str(products_seller),
                        "seller_id": str(products_seller_id),
                        "price": float(products_price),
                        "pick_up": str(products_pick),
                        "delivery_day": str(products_delivery_day),
                        "delivery_price": str(products_delivery_price),
                        "position": str(products_position),
                        "comparison": str(products_comparison)
                    })
            return product_temp
        else:
            for product in temp2["data"]["cards"]:  # по товарам проход
                products_id = int(product['id'])
                products_name = product['title']
                products_url = product['shopLink']
                products_categories = product['category']
                req2 = f"https://kaspi.kz/yml/offer-view/offers/{product['id']}"
                offer_headers = CaseInsensitiveDict()
                offer_headers["Accept"] = "application/json, text/*"
                offer_headers["Referer"] = product["shopLink"]
                offer_headers["Origin"] = "https://kaspi.kz"
                offer_headers["Content-Type"] = "application/json"
                offer_headers["User-Agent"] = agent
                offer_count = 5
                offer_data = json.dumps(
                    {"cityId": "351010000", "id": str(products_id), "merchantUID": "", "limit": offer_count, "page": 0,
                     "sort": True, "baseProductCodes": [], "groups": [], "installationId": "-1"})
                temp3 = requests.post(req2, headers=offer_headers, data=offer_data).json()
                offer_count = temp3['total']
                offer_data = json.dumps(
                    {"cityId": "351010000", "id": str(products_id), "merchantUID": "", "limit": offer_count, "page": 0,
                     "sort": True, "baseProductCodes": [], "groups": [], "installationId": "-1"})
                temp3 = requests.post(req2, headers=offer_headers, data=offer_data).json()
                count = 1
                for offer in temp3["offers"]:
                    products_seller = offer["merchantName"]
                    products_seller_id = offer["merchantId"]
                    products_price = offer["price"]
                    months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь",
                              "Октябрь", "Ноябрь", "Декабрь"]
                    try:
                        products_pick = f"{offer['pickup'][8:10]} {months[int(offer['delivery'][5:7])]}"
                        products_pick = "Есть"
                    except:
                        products_pick = "Нет"
                    try:
                        products_delivery_day = f"{offer['delivery'][8:10]} {months[int(offer['delivery'][5:7])]}"
                        products_delivery_price = "Бесплатно"
                    except:
                        products_delivery_day = "Нет"
                        products_delivery_price = "Бесплатно"
                    products_position = count
                    count += 1
                    a = {
                        "id": int(products_id),
                        "category1": str(products_categories[0]),
                        "category3": str(products_categories[2]),
                        "category2": str(products_categories[1]),
                        "name": str(products_name),
                        "url": str(products_url),
                        "seller": str(products_seller),
                        "seller_id": str(products_seller_id),
                        "price": float(products_price),
                        "pick_up": str(products_pick),
                        "delivery_day": str(products_delivery_day),
                        "delivery_price": str(products_delivery_price),
                        "position": str(products_position),
                        "comparison": str(products_comparison)
                    }
                    print(a["id"])
                    product_temp.append(a)


def compare(merchantid, products):
    count = 2
    randomizer = randint(0, 99)
    ws = wb.active
    ws['A1'] = "Код товара"
    ws['B1'] = "Наименование"
    ws['C1'] = "Категория 1"
    ws['D1'] = "Категория 2"
    ws['E1'] = "Категория 3"
    ws['F1'] = "URL"
    ws['G1'] = "Продавец"
    ws['H1'] = "Цена"
    ws['I1'] = "Дата самовывоза"
    ws['J1'] = "Доставка дата"
    ws['K1'] = "Доставка цена"
    ws['L1'] = "Позиция"
    ws['M1'] = "Сравнение цен"
    products_final = list()
    for product_list in range(len(products)):

        found = False
        save_id = 0
        for product in products[product_list]:
            if str(merchantid) == str(product["seller_id"]):
                save_price = product['price']
                save_id = product['id']
                pos = product['position']
                found = True
                break
        if found:
            for i in range(len(products[product_list])):
                ws[f'A{count}'] = products[product_list][i]['id']
                ws[f'B{count}'] = products[product_list][i]['name']
                try:
                    ws[f'C{count}'] = products[product_list][i]['category1']
                    ws[f'D{count}'] = products[product_list][i]['category2']
                    ws[f'E{count}'] = products[product_list][i]['category3']
                except:
                    ws[f'C{count}'] = ""
                    ws[f'D{count}'] = ""
                    ws[f'E{count}'] = ""
                ws[f'F{count}'] = products[product_list][i]['url']
                ws[f'G{count}'] = products[product_list][i]['seller']
                ws[f'H{count}'] = products[product_list][i]['price']
                ws[f'I{count}'] = products[product_list][i]['pick_up']
                ws[f'J{count}'] = products[product_list][i]['delivery_day']
                ws[f'K{count}'] = products[product_list][i]['delivery_price']
                ws[f'L{count}'] = products[product_list][i]['position']
                # print(products[product_list][i])
                if str(merchantid) == str(products[product_list][i]["seller_id"]) and products[product_list][i][
                    "id"] == save_id:
                    products[product_list][i]['comparison'] = "Это мы"
                    ws[f'M{count}'] = products[product_list][i]['comparison']
                elif products[product_list][i]["id"] == save_id:
                    if products[product_list][i]['price'] < save_price:
                        products[product_list][i]['comparison'] = "Дешевле нас"
                        ws[f'M{count}'] = products[product_list][i]['comparison']
                    elif products[product_list][i]["price"] > save_price:
                        products[product_list][i]['comparison'] = "Дороже нас"
                        ws[f'M{count}'] = products[product_list][i]['comparison']
                    elif products[product_list][i]["price"] == save_price:
                        products[product_list][i]['comparison'] = "Как мы"
                        ws[f'M{count}'] = products[product_list][i]['comparison']
                save_id = products[product_list][i]["id"]
                count += 1
        if not found:
            for i in range(len(products[product_list])):
                if int(products[product_list][i]["position"]) < 2:
                    if save_id != products[product_list][i]["id"]:
                        products[product_list][i]['comparison'] = "Нас нет"
                        ws[f'A{count}'] = products[product_list][i]['id']
                        ws[f'B{count}'] = products[product_list][i]['name']
                        try:
                            ws[f'C{count}'] = products[product_list][i]['category1']
                            ws[f'D{count}'] = products[product_list][i]['category2']
                            ws[f'E{count}'] = products[product_list][i]['category3']
                        except:
                            ws[f'C{count}'] = ""
                            ws[f'D{count}'] = ""
                            ws[f'E{count}'] = ""
                        ws[f'F{count}'] = products[product_list][i]['url']
                        ws[f'G{count}'] = products[product_list][i]['seller']
                        ws[f'H{count}'] = products[product_list][i]['price']
                        ws[f'I{count}'] = products[product_list][i]['pick_up']
                        ws[f'J{count}'] = products[product_list][i]['delivery_day']
                        ws[f'K{count}'] = products[product_list][i]['delivery_price']
                        ws[f'L{count}'] = products[product_list][i]['position']
                        ws[f'M{count}'] = products[product_list][i]['comparison']
                        count += 1
                        break
        # for i in range(len(products[product_list])):
        #     ws[f'A{count }'] = products[product_list][i]['i   d']
        #     ws[f'B{i + 2 + product_list}'] = products[product_list][i]['name']
        #     try:
        #         ws[f'C{i + 2 + product_list}'] = products[product_list][i]['category1']
        #         ws[f'D{i + 2 + product_list}'] = products[product_list][i]['category2']
        #         ws[f'E{i + 2 + product_list}'] = products[product_list][i]['category3']
        #     except:
        #         ws[f'C{i + 2 + product_list}'] = ""
        #         ws[f'D{i + 2 + product_list}'] = ""
        #         ws[f'E{i + 2 + product_list}'] = ""
        #     ws[f'F{i + 2 + product_list}'] = products[product_list][i]['url']
        #     ws[f'G{i + 2 + product_list}'] = products[product_list][i]['seller']
        #     ws[f'H{i + 2 + product_list}'] = products[product_list][i]['price']
        #     ws[f'I{i + 2 + product_list}'] = products[product_list][i]['pick_up']
        #     ws[f'J{i + 2 + product_list}'] = products[product_list][i]['delivery_day']
        #     ws[f'K{i + 2 + product_list}'] = products[product_list][i]['delivery_price']
        #     ws[f'L{i + 2 + product_list}'] = products[product_list][i]['position']
        #     ws[f'M{i + 2 + product_list}'] = products[product_list][i]['comparison']
        # if product["name"] == "Гербор Кентуки Tol Белый":
        # #     print(product_list)
        # product_list_final.append(product_list)
    print("Success comparison")
    # return products_final
    name = f"{category}{randint(0, 99)}.xlsx"
    # price = 0
    # for g in range(len(products)):
    #     for i in range(len(products[g])):
    #         if ws[f'L{i + 2 + g}'] == "Это мы" or ws[f'L{i + 2 + g}'] == "Дороже нас"  or ws[f'L{i + 2 + g}'] == "Дешевле нас" or ws[f'L{i + 2 + g}'] == "Как мы":
    #             price = ws[f'H{i + 2 + g}']
    #         elif price != 0:
    #             if[f'']
    wb.save(f"{category}{randomizer}.xlsx")


def start_excel(category, products):
    # from openpyxl import load_workbook
    #
    # # load excel file
    # workbook = load_workbook(filename="")
    #
    # # open workbook
    # sheet = workbook.active
    #
    # # modify the desired cell
    # sheet["A1"] = "Full Name"
    #
    # # save the file
    # workbook.save(filename="csv/output.xlsx")
    pass


if __name__ == '__main__':
    from tkinter import *
    from tkinter import ttk
    import threading
    import logging

    root = Tk()
    frm = ttk.Frame(root, padding=10)
    frm.grid()

    agent = "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1650.69 Safari/537.36"
    print(agent)
    wb = Workbook()
    list_of_categories = ["Мебель", "Товары для дома и дачи", "Украшения", "Автотовары",
                          "Строительство, ремонт", "Одежда", "Красота и здоровье", "Досуг, книги",
                          "Спорт, туризм", "Аксессуары", "Детские товары", "Обувь", "Аптека", "Бытовая техника",
                          "Телефоны и гаджеты", "Компьютеры", "Продукты питания", "Подарки, товары для праздников",
                          "ТВ, Аудио, Видео", "Канцелярские товары", "Товары для животных"]

    url = StringVar()
    merchant = StringVar()

    ttk.Entry(frm, textvariable=url).grid(column=1, row=0)
    ttk.Label(frm, text="URL").grid(column=0, row=0)
    ttk.Entry(frm, textvariable=merchant).grid(column=1, row=2)
    ttk.Label(frm, text="merchantid").grid(column=0, row=2)

    ttk.Button(frm, text="Search", command=get_a_link).grid(column=1, row=3)
    root.mainloop()


